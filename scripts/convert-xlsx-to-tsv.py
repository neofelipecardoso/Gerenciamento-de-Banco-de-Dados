from pathlib import Path
import pandas as pd
import csv
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

INPUT_DIR = Path("datasets_xlsx")
ENCODING = "utf-8"
LINETERMINATOR = "\n"
OUTPUT_DIR = Path("datasets_tsv")
HEADER_ROWS_TO_SKIP = 3
FOOTER_ROWS_TO_SKIP = 1


def unmerge_and_fill(sheet):
    merged_ranges = list(sheet.merged_cells.ranges)

    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))

        top_left_cell = sheet.cell(row=min_row, column=min_col)
        value = top_left_cell.value

        sheet.unmerge_cells(str(merged_range))

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row=row, column=col).value = value


def extract_content(ws, file):
    all_rows = list(ws.values)
    trimmed_rows = all_rows[HEADER_ROWS_TO_SKIP:-FOOTER_ROWS_TO_SKIP]

    if not trimmed_rows:
        raise ValueError(f"{file} is too small.")

    return trimmed_rows[1:], trimmed_rows[0]


def convert_file(file):
    if not file.suffix.lower() == ".xlsx" or not file.is_file():
        raise ValueError(f"The input directory must contain only XLSX files.: {file}")

    output_file = OUTPUT_DIR / file.name.replace(".xlsx", ".tsv")

    print(f"Converting '{file}' to '{output_file}'")

    ws = load_workbook(file, data_only=True).active

    unmerge_and_fill(ws)

    data, cols = extract_content(ws, file)

    df = pd.DataFrame(data, columns=cols)

    df.to_csv(
        output_file,
        sep="\t",
        index=False,
        encoding=ENCODING,
        lineterminator=LINETERMINATOR,
        quoting=csv.QUOTE_MINIMAL,
    )


def main():
    print("Starting 'convert-xlsx-to-tsv.py'...")

    if not INPUT_DIR.exists():
        raise FileNotFoundError(f"Input directory not found: {INPUT_DIR}")

    files = sorted(INPUT_DIR.glob("*"))

    OUTPUT_DIR.mkdir(exist_ok=True)

    for file in files:
        try:
            convert_file(file)
        except Exception as e:
            raise RuntimeError(f"Error converting '{file}': {e}")

    print("Done")


if __name__ == "__main__":
    main()
